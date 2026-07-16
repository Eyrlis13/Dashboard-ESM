"""
Parser ESM — lit les bilans ergo Excel, anonymise, et produit un dataset propre.
AUCUN nom / adresse / téléphone / mail n'est conservé.
Sortie : liste de dicts, ID = 'DEP-NN' (département + index anonyme).
"""
import openpyxl, glob, os, re, unicodedata, warnings, json, sys
warnings.filterwarnings('ignore')

# --- Chemins portables (aucun chemin en dur) --------------------------------
# Par défaut : dossier "bilans/" à côté du script, sortie "dataset.json" idem.
# Surchargeable via variables d'environnement ou arguments ligne de commande.
HERE = os.path.dirname(os.path.abspath(__file__))
BILANS_DIR = os.environ.get('ESM_BILANS_DIR') or os.path.join(HERE, 'bilans')
OUTPUT_JSON = os.environ.get('ESM_OUTPUT') or os.path.join(HERE, 'dataset.json')

def norm(s):
    if s is None: return ''
    return re.sub(r'\s+',' ',''.join(c for c in unicodedata.normalize('NFD',str(s)) if unicodedata.category(c)!='Mn').lower().strip())

def find_sheet(wb,*kw):
    for name in wb.sheetnames:
        if all(k in norm(name) for k in kw): return wb[name]
    return None

CITY_DEP=[('blois','41'),('joue les tours','37'),('tours','37'),('vendome','41'),
          ('romorantin','41'),('amboise','37'),('contres','41'),('vineuil','41')]

def scan_right(ws, c, maxoff=5):
    """Renvoie la 1re valeur non vide à droite du libellé, sur la même ligne.
    Certains bilans placent la valeur en colonne B, d'autres en C, D ou F :
    on balaie donc plusieurs colonnes au lieu de ne regarder que la voisine."""
    for off in range(1, maxoff+1):
        v = ws.cell(row=c.row, column=c.column+off).value
        if v not in (None, ''):
            return v
    return None

def get_department(wb):
    ws=find_sheet(wb,'infos') or find_sheet(wb,'benef')
    if ws:
        for row in ws.iter_rows(min_row=1,max_row=14):
            for c in row:
                if c.value and 'adresse' in norm(c.value):
                    addr=str(scan_right(ws,c) or '')
                    m=re.search(r'\b(\d{5})\b',addr)
                    if m: return m.group(1)[:2]
    blob=''
    for shname in wb.sheetnames:
        for row in wb[shname].iter_rows():
            for c in row:
                if isinstance(c.value,str): blob+=' '+norm(c.value)+' '
    for city,dep in CITY_DEP:
        if f' {city} ' in blob or f' {city}.' in blob or f' {city},' in blob:
            return dep
    return None

def fesi_score(wb,when):
    ws=find_sheet(wb,'fes',when)
    if ws is None: return None
    for row in ws.iter_rows():
        for c in row:
            if c.value and 'score' in norm(c.value):
                for cc in row:
                    if isinstance(cc.value,(int,float)) and cc.column>c.column:
                        return cc.value if 0<cc.value<=64 else None   # 0/vide = non rempli
    return None

def get_infos(wb):
    import datetime
    ws=find_sheet(wb,'infos') or find_sheet(wb,'benef')
    out={'age':None,'gir':None,'zone':None,'type_res':None,'situ_fam':None,'permis':None}
    if ws is None: return out
    def right(c): return scan_right(ws,c)
    for row in ws.iter_rows(min_row=1,max_row=27):
        for c in row:
            if not c.value: continue
            k=norm(c.value)
            if k.startswith('date de naissance'):
                v=right(c)
                if isinstance(v,datetime.datetime): out['age']=2026-v.year
                elif isinstance(v,str):
                    m=re.search(r'(\d{4})',v)
                    if m: out['age']=2026-int(m.group(1))
            elif k=='gir':
                v=right(c)
                try: out['gir']=int(float(v))
                except: pass
            elif 'zone de residence' in k: out['zone']=str(right(c)).strip() if right(c) else None
            elif 'type de residence' in k: out['type_res']=str(right(c)).strip() if right(c) else None
            elif 'situation familiale' in k: out['situ_fam']=str(right(c)).strip() if right(c) else None
            elif k.startswith('permis'):
                v=right(c); out['permis']=str(v).strip() if v else None
    return out

# Modal shift: count modes used avant vs après, and flag conduite seule / alternatives
CONDUITE_SEULE=['conduit seul','conduite seule']
ALTERNATIVES=['transports en commun','transport en commun','marche avec at','taxi','vtc',
              'transports a la demande','transport a la demande','covoiturage','vsl','ambulance',
              'livraison','velo','scooter','navette','bus','tram','train']

def get_modes(wb,when):
    ws=find_sheet(wb,'modes',when)
    modes=[]
    if ws is None: return modes
    for row in ws.iter_rows(min_row=4,max_row=20):
        for col in ['C','F']:
            v=ws[f'{col}{row[0].row}'].value
            if v and str(v).strip(): modes.append(norm(v))
    return modes

def count_flags(modes):
    seule=sum(1 for m in modes if any(x in m for x in CONDUITE_SEULE))
    alt=sum(1 for m in modes if any(x in m for x in ALTERNATIVES))
    return seule,alt

def get_objectives(wb):
    ws=find_sheet(wb,'bilan','fin')
    res={'total':0,'atteint':0,'partiel':0,'non':0}
    if ws is None: return res
    for row in ws.iter_rows(min_row=4,max_row=8):
        obj=ws[f'B{row[0].row}'].value
        att=ws[f'J{row[0].row}'].value
        if obj and 'objectif' in norm(obj) and len(norm(obj))>norm('objectif 9 :').__len__():
            na=norm(att) if att else ''
            if 'partiel' in na: res['partiel']+=1; res['total']+=1
            elif 'non' in na: res['non']+=1; res['total']+=1
            elif 'atteint' in na: res['atteint']+=1; res['total']+=1
    return res

def get_verbatims(wb):
    out={'beneficiaire':None,'ergo':None}
    ws=find_sheet(wb,'bilan','fin')
    if ws:
        for row in ws.iter_rows():
            for c in row:
                if c.value and 'ressenti de l' in norm(c.value):
                    # text is usually in the next rows
                    txt=ws.cell(row=c.row+1,column=c.column).value
                    if txt and len(str(txt))>15: out['beneficiaire']=str(txt).strip()
                if c.value and 'synthese de l' in norm(c.value):
                    txt=ws.cell(row=c.row+1,column=c.column).value
                    if txt and len(str(txt))>15: out['ergo']=str(txt).strip()
    return out

# ---- Ressentis : 4 échelles avant/après (voir REGLES_RESSENTIS.md) ----
RESS_ITEMKW={'aise':['aise'],'satisfaction':['satisf'],
             'confiance':['crain','chuter','tomber','peur'],'serenite':['serein']}
RESS_LIKERT={'pas du tout':1,'peu':2,'assez':3,"a l'aise":4,'a l aise':4,'satisfait':4,
             'tres serein':4,'tres a l':4,'constamment':1,'souvent':2,'quelques fois':3,'jamais':4}

def _ress_marked(c):
    try:
        f=c.fill
        if f and f.patternType and f.patternType!='none':
            rgb=getattr(f.fgColor,'rgb',None)
            if isinstance(rgb,str) and rgb not in ('FFFFFFFF','00000000'): return True
            if getattr(f.fgColor,'theme',None) is not None and f.patternType=='solid': return True
    except: pass
    return False

def _ress_bold(c):
    try: return bool(c.font and c.font.bold)
    except: return False

def _ress_fontcolored(c):
    """Vrai si la police a une couleur RGB explicite (≠ noir/défaut) — certains
    bilans marquent la réponse choisie en rouge sur l'échelle 1→10."""
    try:
        col=c.font.color
        rgb=getattr(col,'rgb',None) if col is not None else None
        if isinstance(rgb,str) and rgb not in ('FF000000','00000000','FFFFFFFF'):
            return True
    except: pass
    return False

def _ress_positivity(item,fmt,val):
    """Convertit une réponse en score 0-10 « plus haut = mieux » (confiance = crainte inversée).
    'pct' = déjà 0-10 (likert positionnel, ancre gauche = pire) ; 'scale10' = position 1-10 ;
    'num' = nombre 0-10 (confiance inversée)."""
    if val is None or fmt is None: return None
    if fmt=='scale10':  p=(val-1)/9*10
    elif fmt=='pct':    p=val
    elif fmt=='num':    p=val if item!='confiance' else 10-val
    else: return None
    return round(max(0.0,min(10.0,p)),1)

# Mots des ancres qui identifient LA ligne de réponse propre à chaque item
RESS_ANCHORWORDS={'aise':['aise'],'satisfaction':['satisf'],
                  'confiance':['constamment','souvent','quelques fois','jamais'],
                  'serenite':['serein']}

def _ress_answer_on_row(ws,r,is_scale):
    """Réponse marquée (gras / couleur police / surlignage) sur la ligne r.
    Échelle 1-10 → ('scale10', chiffre). 4 niveaux qualitatifs → ('pct', score 0-10)
    calculé par la POSITION de la case cochée (indépendant du vocabulaire des ancres)."""
    if is_scale:
        for c in ws[r]:
            if not (2<c.column<=12): continue
            if isinstance(c.value,(int,float)) and not isinstance(c.value,bool):
                if _ress_bold(c) or _ress_fontcolored(c) or _ress_marked(c):
                    return ('scale10',c.value)
        return None
    # likert : rang de l'ancre marquée parmi les ancres (gauche = pire = 0)
    anchors=[c for c in ws[r] if 2<c.column<=12 and isinstance(c.value,str) and c.value.strip()]
    anchors.sort(key=lambda c:c.column)
    n=len(anchors)
    if n<2: return None
    for i,c in enumerate(anchors):
        if _ress_marked(c) or _ress_fontcolored(c):
            return ('pct', i/(n-1)*10)
    return None

def _ress_sheet(ws):
    # localiser les 4 questions (colonnes A/B)
    qrows={}
    for row in ws.iter_rows():
        for c in row:
            if c.column>2 or not isinstance(c.value,str) or len(c.value)<12: continue
            n=norm(c.value)
            for item,kws in RESS_ITEMKW.items():
                if item not in qrows and any(k in n for k in kws):
                    qrows[item]=c.row
    others=set(qrows.values())
    res={}
    for item,qrow in qrows.items():
        ans=None; lone=None
        # la réponse peut être sur la ligne de la question, juste en dessous ou juste au-dessus
        for r in (qrow,qrow+1,qrow-1,qrow+2):
            if r<1 or r>ws.max_row: continue
            if r!=qrow and r in others: continue        # ne pas empiéter sur une autre question
            cells=[c for c in ws[r] if 2<c.column<=12]
            nums=[c for c in cells if isinstance(c.value,(int,float)) and not isinstance(c.value,bool)]
            is_scale=len(nums)>=8
            if is_scale:
                a=_ress_answer_on_row(ws,r,True)
                if a: ans=a; break
            else:
                txt=' '.join(norm(c.value) for c in cells if isinstance(c.value,str))
                if any(w in txt for w in RESS_ANCHORWORDS[item]):   # ligne d'ancres de CET item
                    a=_ress_answer_on_row(ws,r,False)
                    if a: ans=a; break
                elif len(nums)==1 and lone is None:
                    lone=nums[0].value                              # nombre isolé (secours)
        if ans: res[item]=_ress_positivity(item,*ans)
        elif lone is not None: res[item]=_ress_positivity(item,'num',lone)
        else: res[item]=None
    return res

def get_ressentis(wb):
    out={d:{'avant':None,'apres':None} for d in RESS_ITEMKW}
    for when in ('avant','apres'):
        ws=find_sheet(wb,'ressentis',when)
        if ws is None: continue
        vals=_ress_sheet(ws)
        for d in RESS_ITEMKW: out[d][when]=vals.get(d)
    return out

def load_at_curated():
    """Classification curée des aides techniques (at_curated.csv, non versionné) :
    statut 'mise_en_place' / 'essayee' / 'preconisee', par dossier."""
    import csv
    path=os.path.join(HERE,'at_curated.csv')
    at={}
    if not os.path.isfile(path): return at
    with open(path,encoding='utf-8-sig') as fh:
        for row in csv.DictReader(fh):
            cle=(row.get('cle') or '').strip().upper()
            typ=(row.get('type') or '').strip()
            st=(row.get('statut') or '').strip()
            if cle and typ and st: at.setdefault(cle,[]).append({'type':typ,'statut':st})
    return at

def load_manual_ressentis():
    """Saisies manuelles (ressentis_manuel.csv, non versionné). Priment sur l'auto."""
    import csv
    path=os.path.join(HERE,'ressentis_manuel.csv')
    man={}
    if not os.path.isfile(path): return man
    with open(path,encoding='utf-8-sig') as fh:
        for row in csv.DictReader(fh):
            cle=(row.get('cle') or '').strip().upper()
            if cle: man[cle]={k:v for k,v in row.items() if k!='cle'}
    return man

# ---- Déplacements détaillés : modes + types de sorties (grille "Modes de déplacements") ----
MODE_CATS=['auto','marche','accompagne','alternatif','domicile']
def mode_cat(v):
    if not v: return None
    n=norm(v)
    if 'conduit seul' in n or 'conduite seul' in n: return 'auto'
    if 'marche' in n: return 'marche'
    if 'proche' in n or 'covoiturage' in n: return 'accompagne'
    if 'livraison' in n or 'domicile' in n: return 'domicile'
    if any(k in n for k in ['transport','commun','demande','taxi','vtc','vsl','ambulance',
                            'bus','tram','navette','trottinette','minibus','partag','scooter']): return 'alternatif'
    if 'autre' in n: return 'alternatif'
    return None

def trip_cat(v):
    n=norm(v)
    if any(k in n for k in ['course','shopping','pharmacie','magasin']): return 'courses'
    if any(k in n for k in ['rdv','medic','special','traitant','veterinaire','sante']): return 'sante'
    if any(k in n for k in ['loisir','promener','benevol']): return 'loisirs'
    if 'famille' in n or 'amis' in n: return 'social'
    if 'administ' in n or 'demarche' in n: return 'demarches'
    return 'autres'

def get_deplacements(wb):
    out={}
    for when in ('avant','apres'):
        ws=find_sheet(wb,'modes',when)
        modes={c:0 for c in MODE_CATS}
        sorties={}; pairs=[]
        if ws is not None:
            for r in range(5,20):
                b=ws.cell(row=r,column=2).value
                if not (b and str(b).strip()): continue
                tcat=trip_cat(str(b))
                mcats=[]
                for col in (3,6):                       # mode principal (C) + secondaire (F)
                    mc=mode_cat(ws.cell(row=r,column=col).value)
                    if mc: mcats.append(mc); modes[mc]+=1
                if mcats:
                    sorties[tcat]=sorties.get(tcat,0)+1
                    pairs.append([tcat,mcats[0]])       # (type de sortie, mode principal)
        out[when]={'modes':modes,'sorties':sorties,'pairs':pairs}
    return out

def build_dataset(bilans_dir=None):
    bilans_dir = bilans_dir or BILANS_DIR
    files=sorted(glob.glob(os.path.join(bilans_dir, 'Bilan_ESM_Ergo_*.xlsx')))
    def bn(f):
        b=os.path.basename(f).replace('Bilan_ESM_Ergo_','').replace('.xlsx','')
        b=b.replace('VF_','').replace('bis_V3_','').replace('Mme_','').replace('_1','')
        b=re.sub(r'_?termin[eé]?_?$','',b,flags=re.IGNORECASE)   # suffixe "_termin_"/"terminé"
        return b.strip().upper().replace(' ','').strip('_')
    groups={}
    for f in files: groups.setdefault(bn(f),[]).append(f)
    canon={k:([x for x in v if 'termin' in x] or v)[0] for k,v in groups.items()}

    records=[]
    seq=0
    manual_ress=load_manual_ressentis()
    at_curated=load_at_curated()
    for k,f in sorted(canon.items()):
        wb=openpyxl.load_workbook(f,data_only=True)
        dep=get_department(wb) or 'NR'
        seq+=1
        anon_id=str(seq)          # appellation simple : 1, 2, 3 … (aucun nom, aucun code département)
        info=get_infos(wb)
        real_names=get_names_from_file(wb)
        # the participant key (from filename) IS the surname — always scrub it + accented variants
        real_names.add(k.capitalize()); real_names.add(k)
        real_names.add(k.replace('THEOPHILE','Théophile'))
        fa,fp=fesi_score(wb,'avant'),fesi_score(wb,'apres')
        # GUERAULT : FES-I après non mesuré -> considéré identique à l'avant (stable), décision métier
        if k.upper()=='GUERAULT' and fp is None and fa is not None: fp=fa
        mav,map_=get_modes(wb,'avant'),get_modes(wb,'apres')
        sav,aav=count_flags(mav); sap,aap=count_flags(map_)
        rec={
            'id':anon_id,'departement':dep,
            'age':info['age'],'gir':info['gir'],'zone':info['zone'],
            'type_res':info['type_res'],'permis':info['permis'],
            'fesi_avant':fa,'fesi_apres':fp,
            'fesi_delta':(fp-fa) if (fa is not None and fp is not None) else None,
            'conduite_seule_avant':sav,'conduite_seule_apres':sap,
            'alternatives_avant':aav,'alternatives_apres':aap,
            'objectifs':get_objectives(wb),
        }
        vb=get_verbatims(wb)
        rec['verbatims']={
            'beneficiaire':scrub_text(vb['beneficiaire'],real_names),
            'ergo':scrub_text(vb['ergo'],real_names),
        }
        rec['sexe']=infer_sexe(rec)
        rec['deplacements']=get_deplacements(wb)
        rec['at']=at_curated.get(k.upper(), [])
        # Ressentis (auto) + fusion des saisies manuelles (qui priment)
        rec['ressentis']=get_ressentis(wb)
        mm=manual_ress.get(k.upper())
        if mm:
            for d in RESS_ITEMKW:
                for when in ('avant','apres'):
                    v=mm.get(f"{d}_{when}")
                    if v not in (None,''):
                        try: rec['ressentis'][d][when]=round(float(str(v).replace(',','.')),1)
                        except: pass
        records.append(rec)
    return records


def infer_sexe(rec):
    """Déduit le sexe du bénéficiaire à partir des civilités et accords présents
    dans les verbatims (les civilités « Mme / M. » ne sont pas anonymisées).
    Approche « au mieux » : renvoie 'F', 'H' ou None si indéterminé."""
    vb = rec.get('verbatims') or {}
    txt = ' '.join(filter(None, [vb.get('beneficiaire') or '', vb.get('ergo') or '']))
    low = txt.lower()
    h = 3*len(re.findall(r'\bM\.\s|\bMonsieur\b', txt)) \
        + 2*len(re.findall(r'\bhomme\b|\bmonsieur\b', low)) \
        + len(re.findall(r'\bil\b|\blui-même\b', low))
    f = 3*len(re.findall(r'\bMme\b|\bMadame\b', txt)) \
        + 2*len(re.findall(r'\bdame\b|\bépouse\b', low)) \
        + len(re.findall(r'\belle\b|satisfaite|libérée|rassurée|contente|angoiss?ée|âgée|accompagnée|amenée', low))
    return 'H' if h > f else ('F' if f > h else None)


def get_names_from_file(wb):
    """Récupère le nom réel pour pouvoir le masquer dans les textes. Jamais stocké en sortie."""
    names=set()
    ws=find_sheet(wb,'infos') or find_sheet(wb,'benef')
    if ws:
        for row in ws.iter_rows(min_row=1,max_row=6):
            for c in row:
                if c.value and 'nom' in norm(c.value) and 'prenom' in norm(c.value):
                    full=ws.cell(row=c.row,column=c.column+1).value
                    if full:
                        for tok in re.split(r'[ ,]+',str(full)):
                            if len(tok)>=3: names.add(tok)
    return names

def scrub_text(txt, names):
    if not txt: return txt
    out=txt
    # mask known surname/firstname tokens (case-insensitive, word boundary)
    for n in sorted(names,key=len,reverse=True):
        out=re.sub(rf'\b{re.escape(n)}\b','[…]',out,flags=re.IGNORECASE)
        # also capitalized variants
        out=re.sub(rf'\b{re.escape(n.capitalize())}\b','[…]',out)
    # mask "Mme X" / "M. X" / "Mr X" patterns where X is a Capitalized word
    out=re.sub(r'\b(Mme|M\.|Mr|Monsieur|Madame)\s+[A-ZÉÈÀ][a-zéèàêâîôûç]+','\\1 […]',out)
    # mask soignants/tiers nommés : "Dr X", "Docteur X", "Pr X", "Professeur X"
    out=re.sub(r'\b(Dr|Docteur|Pr|Professeur)\.?\s+[A-ZÉÈÀ][a-zéèàêâîôûç]+',r'\1 […]',out)
    # mask emails / phones just in case
    out=re.sub(r'\b[\w.+-]+@[\w-]+\.[\w.-]+\b','[email]',out)
    out=re.sub(r'\b0\d(?:[ .]?\d{2}){4}\b','[tél]',out)
    return out


if __name__=='__main__':
    # Usage : python3 parse_esm.py [dossier_bilans] [sortie.json]
    bilans_dir = sys.argv[1] if len(sys.argv) > 1 else BILANS_DIR
    out_path   = sys.argv[2] if len(sys.argv) > 2 else OUTPUT_JSON

    if not os.path.isdir(bilans_dir):
        print(f"⚠  Dossier des bilans introuvable : {bilans_dir}")
        print(f"   Crée-le et dépose-y les fichiers Bilan_ESM_Ergo_*.xlsx, puis relance.")
        sys.exit(2)

    data=build_dataset(bilans_dir)
    if not data:
        print(f"⚠  Aucun fichier Bilan_ESM_Ergo_*.xlsx trouvé dans : {bilans_dir}")
        sys.exit(2)

    print(f"{len(data)} participants anonymisés\n")
    for r in data:
        print(f"{r['id']:7s} age={str(r['age']):4s} gir={str(r['gir']):3s} zone={str(r['zone'])[:18]:18s} "
              f"FES-I {str(r['fesi_avant']):3s}->{str(r['fesi_apres']):3s} (Δ{r['fesi_delta']}) "
              f"cond.seule {r['conduite_seule_avant']}->{r['conduite_seule_apres']} "
              f"obj {r['objectifs']['atteint']}A/{r['objectifs']['partiel']}P/{r['objectifs']['non']}N")

    os.makedirs(os.path.dirname(out_path) or '.', exist_ok=True)
    with open(out_path,'w',encoding='utf-8') as fh:
        json.dump(data,fh,ensure_ascii=False,indent=2)
    print(f"\n-> {out_path} écrit (anonymisé)")


# ---- Passe d'anonymisation du TEXTE LIBRE (verbatims / synthèses) ----